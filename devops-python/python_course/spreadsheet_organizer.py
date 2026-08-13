from openpyxl import load_workbook, Workbook

def organize_spreadsheet(file_path):
    # Load the workbook
    workbook = load_workbook(filename=file_path)
    
    # Select the active worksheet
    sheet = workbook.active
    
    # Create a dictionary to hold organized data
    organized_data = {}
    
    # Iterate through the rows in the worksheet
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        category = row[0]  # Assuming the first column is the category
        item = row[1]      # Assuming the second column is the item
        
        if category not in organized_data:
            organized_data[category] = []
        
        organized_data[category].append(item)
    
    # Create a new workbook to save organized data
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    
    # Write organized data to the new sheet
    row_index = 1
    for category, items in organized_data.items():
        new_sheet.cell(row=row_index, column=1, value=category)
        for item in items:
            row_index += 1
            new_sheet.cell(row=row_index, column=2, value=item)
        row_index += 1  # Add an empty row after each category
    
    # Save the new workbook
    new_file_path = file_path.replace('.xlsx', '_organized.xlsx')
    new_workbook.save(new_file_path)
    
    print(f"Organized spreadsheet saved as: {new_file_path}")

organize_spreadsheet('test_data.xlsx')